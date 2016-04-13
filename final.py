#!/usr/bin/env python3
# encoding=utf8

# 本程序是一条龙服务
# 制定存放所有excel数据的目录
# 然后等待就对了


from openpyxl import load_workbook
import sqlite3
import os
import sys
import nltk
import math


if len(sys.argv) != 2:
    print('%s data_location' % sys.argv[0])
    sys.exit(0)

DATA_LOCATION = sys.argv[1]
SQLITE3_FILE = 'final.sqlite3'

try:
    os.unlink(SQLITE3_FILE)
except:
    pass

SQLITE3 = sqlite3.connect(SQLITE3_FILE)


print('Be patient......')

cur = SQLITE3.cursor()
cur.execute(
    'create table data(\
        pdf text,\
        author1 text,\
        author2 text,\
        title text,\
        journal text,\
        page text,\
        year text,\
        info1 text,\
        info2 text,\
        state_year text,\
        comment text,\
        addr text,\
        content text,\
        type text\
        )')
SQLITE3.commit()
cur.close()

#######################################################
cur = SQLITE3.cursor()
total = 0
print('Converting .xlsx files to sqlite3 format')
for parent, dirnames, filenames in os.walk(DATA_LOCATION):
    for filename in filenames:
        if not filename.endswith('.xlsx'):
            continue    # skip
        path = parent + '/' + filename
        wb = load_workbook(path)
        ws = wb.active
        row = 1
        while ws['A' + str(row)].value:
            rid = str(row)
            cur.execute('insert into data values\
                (?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                        (ws['A' + rid].value, ws['B' + rid].value,
                         ws['C' + rid].value, ws['D' + rid].value,
                         ws['E' + rid].value, ws['F' + rid].value,
                         ws['G' + rid].value, ws['H' + rid].value,
                         ws['I' + rid].value, ws['J' + rid].value,
                         ws['K' + rid].value, ws['L' + rid].value,
                         ws['M' + rid].value, ws['N' + rid].value))
            row += 1
            # if row > 3:
            #    break
        total += row-1
        print(filename, row)
SQLITE3.commit()
cur.close()
print('All .xlsx files are converted successfully. Total: %s' % total)
###########################################################

print('----------------------------------------------------------')
print('Extracting words......')

wnl = nltk.stem.WordNetLemmatizer()


def isplural(word):
    lemma = wnl.lemmatize(word, 'n')
    plural = True if word is not lemma else False
    return plural, lemma


def all_ascii(text):
    for c in text:
        if not (c.isalpha() or c == ' ' or c == '-'):
            return False
    return True


def has_upper(text):
    for c in text:
        if c.isupper():
            return True
    return False


def strize(e):
    if type(e) != nltk.tree.Tree:
        if e[1] == 'NN' and has_upper(e[0]):
            return None
        if e[0].startswith('-') or len(e[0]) <= 1:
            return None
        return e[0]
    ch = []
    for i in e.leaves():
        t = strize(i)
        if t is None:
            return None
        ch.append(t)
    return ' '.join(ch)

grammar = """
NOUN:{<NN>} # 名词
NOUN_NOUN:{<NOUN><NOUN>} # 名词+名词
ADJ:{<JJ>} # 形容词
ADJ_NOUN:{<ADJ><ADJ><NOUN>}
{<ADJ><NOUN><NOUN>}
{<ADJ><NOUN>} # 形容词+名词
NOUN_PREP_NOUN:{<ADJ><NOUN><IN><ADJ><NOUN>}
{<NOUN><IN><ADJ><NOUN>}
{<ADJ><NOUN><IN><NOUN>}
{<NOUN><IN><NOUN>} # 名词+介词+名词
ADV_ADJ:{<RB><ADJ>} # 副词+形容词
ADJ_PREP_NOUN:{<ADJ><IN><NOUN>} # 形容词+介词+名词
"""

cp = nltk.RegexpParser(grammar)


def fortree(tree, cur, filename):
    label = tree.label()
    if label == 'NOUN' or label == 'NOUN_NOUN' or label == 'ADJ_NOUN'\
       or label == 'NOUN_PREP_NOUN' or label == 'ADV_ADJ'\
       or label == 'ADJ_PREP_NOUN' or label == 'ADJ':
        text = strize(tree)
        if text is None:
            return
        if all_ascii(text):
            if label == "NOUN_NOUN":
                pass
            elif label == "NOUN":
                if has_upper(text):
                    return
                plural, text = isplural(text)
            elif label == "ADJ_NOUN":
                pass
            elif label == "NOUN_PREP_NOUN":
                pass
            elif label == "ADV_ADJ":
                pass
            elif label == "ADJ_PREP_NOUN":
                pass
            elif label == "ADJ":
                pass
            text = text.lower()
            cur.execute(
                'insert into phrase values(?,?,?)',
                (filename, text, label))
            return
    if tree.height() > 2:
        first = 0
        for subtree in tree.subtrees():
            if first == 0:
                first = 1
            else:
                fortree(subtree, cur, filename)

cur = SQLITE3.cursor()
cur.execute('select content,pdf from data')
rows = cur.fetchall()

cur.execute('create table phrase(pdf text, phrase text, label text)')
SQLITE3.commit()
cur.close()

cur = SQLITE3.cursor()
total = len(rows)
step = 1
for row in rows:
    print('Extracting words - %s/%s' % (step, total))
    step += 1
    text = row[0]
    filename = row[1]
    if text is None:
        continue
    sentences = nltk.sent_tokenize(text)
    for sentence in sentences:
        sentence = nltk.pos_tag(nltk.word_tokenize(sentence))
        tree = cp.parse(sentence)
        fortree(tree, cur, filename)

SQLITE3.commit()
cur.close()

print('\nExtracting words finish!')

########################################
print('--------------------------------------')
print('I\'m a math genius! Wa Ha Ha Ha......')

cur = SQLITE3.cursor()
cur.execute('create table frequence(pdf text,phrase text,label text,'
            'N integer,M integer,frequence double,'
            'frequence_all double,rate double,'
            'TF double, IDF double,TF_IDF double)')
SQLITE3.commit()
cur.close()

cur = SQLITE3.cursor()
cur.execute('select count(phrase) from phrase')
rows = cur.fetchall()

total_phrase = rows[0][0]  # 总词数量

cur.execute('select pdf from phrase group by pdf')
rows = cur.fetchall()

N = len(rows)   # pdf文件总数

cur.execute('select phrase,count(phrase),label from phrase '
            'group by phrase,label')
rows = cur.fetchall()
total_phrases = {}
for row in rows:
    total_phrases[row[0]+ row[2]] = row[1]  # 在所有文件中的，单词词频

cur.execute('select pdf,count(phrase) from phrase group by pdf')
rows = cur.fetchall()
cur.close()

filenames = rows

cursor = SQLITE3.cursor()
total = len(filenames)
step = 1
for filename, phrase_count in filenames:    # pdf名 以及 每个pdf中的词数
    print('Calculating - %s/%s' % (step, total))
    step += 1
    # 在一个文件中(phrase,label)的个数
    cursor.execute('select phrase, count(phrase), label'
                   ' from phrase where pdf = ? group by phrase, label',
                   (filename,))
    phrases = []
    rows = cursor.fetchall()
    for row in rows:
        phrases.append((row[0], row[1], row[2]))

    for phrase in phrases:
        # 方法B
        frequence = phrase[1] / phrase_count

        text = phrase[0] + phrase[2]
        frequence_all = total_phrases[text] / total_phrase
        rate = math.log(frequence / frequence_all, 10)

        # 方法C
        TF = frequence
        cursor.execute('select pdf from phrase where '
                       'phrase=? and label=? group by pdf',
                       (phrase[0], phrase[2]))
        rows = cursor.fetchall()
        PN = len(rows)
        IDF = math.log(N / PN, 10)
        TF_IDF = TF * IDF

        cursor.execute('insert into frequence values(?,?,?,?,?,?,?,?,?,?,?)',
                       (filename, phrase[0], phrase[2], phrase[1],
                        total_phrases[text], frequence, frequence_all,
                        rate, TF, IDF, TF_IDF))


SQLITE3.commit()
cursor.close()

print('\n==============================')
print('Done!!! Enjoy the Big Data')

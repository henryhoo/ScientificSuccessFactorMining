#!/usr/bin/env python3
# encoding=utf8
# 本程序读取制定的主题词文件，以及data表，然后...生成sqlite3文件
#
#
# 3. 我会给你大类的主题词表，每个主题词下面有对应的关键词。比如new:(new method),(new idea)....  然后这些主题词将会和时间，发布# 时间，类别做统计。得到三张二维表。主题词表，year表，state——year表，type表的样例见附件excle中的sheet1，sheet2，sheet3，sheet4。
# 【主题词表列说明】：A  主题词
#                    B   关键词
#                    C   关键词在所有PDF中的词频（即之前frequent里面的M）
#                    D   总体相关度=关键词a的词频 /  所有关键词的词频数之和  （比如样例中该值为 列C中的词频/2805）
#                    E  类别内相关度=关键词a的词频 / a所属类别中的所有关键词词频数之和    （比如样例中first类中词的该值为 列C中的词频/837）


import sqlite3
import openpyxl
import sys
import os


if len(sys.argv) != 3:
    print('Usage: %s topic.xlsx final.sqlite3' % sys.argv[0])
    sys.exit(1)


EXCEL_BOOK = sys.argv[1]
DATA_FILE = sys.argv[2]
TOPIC_FILE = 'topic.sqlite3'

try:
    os.unlink(TOPIC_FILE)
except:
    pass


TOPIC_DB = sqlite3.connect(TOPIC_FILE)
tcur = TOPIC_DB.cursor()

DATA_DB = sqlite3.connect(DATA_FILE)
dcur = DATA_DB.cursor()
# topic 主题词
# phrase 关键字
# pdf 所在pdf文件名
# year 论文发表的年份
# state_year 陈述发表的年份
# N 该关键字在该pdf中出现的次数
tcur.execute('create table basic(topic text, phrase text, '
             'pdf text, year text, state_year text,type text, N integer, '
             'addr text, matchword text, country_code text, '
             'country_name text, continent text, region text)')
TOPIC_DB.commit()

wb = openpyxl.load_workbook(EXCEL_BOOK)
ws = wb.active
i = 1
while ws['A' + str(i)].value:
    r = str(i)
    i += 1
    topic = ws['A' + r].value
    phrase = ws['B' + r].value
    dcur.execute('select data.pdf,year,state_year,type,count(phrase.phrase)'
                 ' as N,pdf_country.addr,pdf_country.matchword,'
                 'pdf_country.country_code,pdf_country.country_name,'
                 'pdf_country.continent,pdf_country.region from data,phrase,'
                 'pdf_country where phrase.phrase=? '
                 'and data.pdf = phrase.pdf and pdf_country.pdf=data.pdf '
                 'group by phrase.phrase, data.pdf;', (phrase,))
    rows = dcur.fetchall()
    for row in rows:
        tcur.execute('insert into basic values(?,?,?,?,?,?,?,?,?,?,?,?,?)',
                     (topic, phrase, row[0], row[1], row[2], row[3], row[4],
                      row[5], row[6], row[7], row[8], row[9], row[10]))
TOPIC_DB.commit()
dcur.close()


# topic
tcur.execute('create view topic as select topic,phrase,sum(N) as M, '
             'CAST(sum(N) as float)/'
             '(select sum(N) from basic) as P, CAST(sum(N) as float)/'
             '(select sum(N) from basic where topic = B.topic) as Q '
             'from basic as B group by topic,phrase')

# state_year
tcur.execute('create view state_year_detail as select topic, phrase, '
             'state_year, sum(N) as N from basic group by state_year, '
             'phrase,topic')
tcur.execute('create view state_year as select topic,state_year, '
             'sum(N) as M from state_year_detail group by topic,state_year')

# continent
tcur.execute('create view continent_detail as '
             'select topic,phrase,continent,sum(N) as N from basic '
             'group by continent,phrase,topic')
tcur.execute('create view continent as select topic,continent,sum(N) as M '
             'from continent_detail group by topic,continent')


# region
tcur.execute('create view region_detail as '
             'select topic,phrase,region,sum(N) as N from basic '
             'group by region,phrase,topic')
tcur.execute('create view region as select topic,region,sum(N) as M '
             'from region_detail group by topic,region')

# year
tcur.execute('create view year_detail as select topic, phrase, year,sum(N) '
             'as N from basic group by year, phrase,topic')
tcur.execute('create view year as select topic,year, '
             'sum(N) as M from year_detail group by topic,year')
TOPIC_DB.commit()

# type
tcur.execute('create table type_basic(topic text, phrase text, type text, '
             'N integer)')
TOPIC_DB.commit()

tcur.execute('select topic, phrase, type, N from basic')
rows = tcur.fetchall()
for row in rows:
    topic = row[0]
    phrase = row[1]
    types = row[2].split('*')
    N = row[3]
    for t in types:
        tcur.execute('insert into type_basic values(?,?,?,?)',
                     (topic, phrase, t, N))
TOPIC_DB.commit()

tcur.execute('create view type_detail as select topic,phrase,type,sum(N) as N '
             'from type_basic group by topic,phrase,type')
tcur.execute('create view type as select topic,type, sum(N) as M '
             'from type_basic group by topic,type;')
TOPIC_DB.commit()
tcur.close()


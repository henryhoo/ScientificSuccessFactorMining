#!/usr/bin/env python3
# encoding=utf8
# 将xlsx中的pdf_country转化为sqlite3

import sqlite3
import sys
from openpyxl import load_workbook


if len(sys.argv) != 3:
    print('%s final.sqlite3 pdf_country.xlsx' % sys.argv[0])
    sys.exit(0)

SQLITE3_DB = sqlite3.connect(sys.argv[1])
wb = load_workbook(sys.argv[2])

cur = SQLITE3_DB.cursor()

cur.execute('drop table if exists pdf_country')
cur.execute('CREATE TABLE pdf_country(pdf text, addr text, matchword text, '
            'country_code text, country_name text, continent text, '
            'region text)')

ws = wb.active


i = 1
while ws['A'+str(i)].value:
    pdf = ws['A'+str(i)].value
    addr = ws['B'+str(i)].value
    matchword = ws['C'+str(i)].value
    country_code = ws['D'+str(i)].value
    country_name = ws['E'+str(i)].value
    continent = ws['F'+str(i)].value
    region = ws['G'+str(i)].value
    cur.execute('insert into pdf_country(pdf,addr,matchword,country_code,'
                'country_name,continent,region) values(?,?,?,?,?,?,?)',
                (pdf, addr, matchword, country_code,
                 country_name, continent, region))
    i += 1

SQLITE3_DB.commit()

#!/usr/bin/env python3
# encoding=utf8

from openpyxl import load_workbook
import sqlite3
import sys
import re

if len(sys.argv) != 3:
    print('%s final.sqlite3 country.xlsx' % sys.argv[0])
    sys.exit(0)

SQLITE3_DB = sqlite3.connect(sys.argv[1])
cur = SQLITE3_DB.cursor()
cur.execute('drop table if exists city')
cur.execute('drop table if exists country')
cur.execute('drop table if exists pdf_country')

cur.execute('create table city(name text, country_code text, district text,'
            'population int unsigned)')

wb = load_workbook(sys.argv[2])
ws = wb['city']
i = 1
while ws['A'+str(i)].value:
    name = ws['A'+str(i)].value
    country_code = ws['B'+str(i)].value
    district = ws['C'+str(i)].value
    population = ws['D'+str(i)].value
    i += 1

    cur.execute('insert into city(name,country_code,district,population) '
                'values(?,?,?,?)', (name, country_code, district, population))

cur.execute('create table country(code text,name text,continent text,'
            'region text,surface_area text,indep_year text,'
            'populcation int unsigned,life_expectancy text, '
            'gnp text, gnpold text, local_name text, government_form text,'
            'head_of_state text,capital text, code2 text)')
ws = wb['country']
i = 1
while ws['A'+str(i)].value:
    col1 = ws['A'+str(i)].value
    col2 = ws['B'+str(i)].value
    col3 = ws['C'+str(i)].value
    col4 = ws['D'+str(i)].value
    col5 = ws['E'+str(i)].value
    col6 = ws['F'+str(i)].value
    col7 = ws['G'+str(i)].value
    col8 = ws['H'+str(i)].value
    col9 = ws['I'+str(i)].value
    col10 = ws['J'+str(i)].value
    col11 = ws['K'+str(i)].value
    col12 = ws['L'+str(i)].value
    col13 = ws['M'+str(i)].value
    col14 = ws['N'+str(i)].value
    col15 = ws['O'+str(i)].value
    i += 1
    cur.execute('insert into country values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                (col1, col2, col3, col4, col5, col6, col7, col8, col9, col10,
                 col11, col12, col13, col14, col15))

cur.execute('create table pdf_country(pdf text, addr text, matchword text, '
            'country_code text, country_name text, '
            'continent text, region text)')

cur.execute('select pdf, addr from data')
pdfs = []
rows = cur.fetchall()
for row in rows:
    pdfs.append({'pdf': row[0], 'addr': row[1]})

cur.execute('select city.name, country.code, country.name, country.continent, '
            'country.region, city.district from city,country where '
            'city.country_code==country.code')
cities = []
rows = cur.fetchall()
for row in rows:
    district = row[5]
    if district is None:
        district = ''
    cities.append({'city_name': row[0], 'country_code': row[1],
                   'country_name': row[2], 'continent': row[3],
                   'region': row[4], 'district': district})


def pre_addr(addr):
    if not addr:
        return ''
    return ' ' + addr.replace(',', ' ').replace('.', '').replace(';', '') + ' '


def pre_matchword(word):
    word = word.replace(' ', '\\s+')
    return '\\s+' + word + '\\s+'

country_cities = sorted(cities, key=lambda X: -len(X['country_name']))
city_cities = sorted(cities, key=lambda X: -len(X['city_name']))
district_cities = sorted(cities, key=lambda X: -len(X['district']))

i = 1
total = len(pdfs)
for data in pdfs:
    print('%s/%s' % (i, total))
    i += 1
    pdf = data['pdf']
    addr = pre_addr(data['addr'])
    match = None

    for city in country_cities:
        try:
            matchword = pre_matchword(city['country_name'])
            p = re.compile(matchword, re.IGNORECASE)
            if p.search(addr):
                match = city
                match['matchword'] = city['country_name']
                break
        except Exception as e:
            pass

    if match is None:
        for city in city_cities:
            try:
                matchword = pre_matchword(city['city_name'])
                p = re.compile(matchword, re.IGNORECASE)
                if p.search(addr):
                    match = city
                    match['matchword'] = city['city_name']
                    break
            except Exception as e:
                pass

    if match is None:
        for city in district_cities:
            try:
                matchword = pre_matchword(city['district'])
                p = re.compile(matchword, re.IGNORECASE)
                if p.search(addr):
                    match = city
                    match['matchword'] = city['district']
                    break
            except Exception as e:
                pass

    if match is None:
        match = {'city_name': '', 'country_code': '',
                 'country_name': '', 'continent': '',
                 'region': '', 'district': '', 'matchword': ''}

    cur.execute('insert into pdf_country(pdf,addr,matchword,country_code,'
                'country_name,continent,region) values(?,?,?,?,?,?,?)',
                (data['pdf'], data['addr'], match['matchword'],
                 match['country_code'], match['country_name'],
                 match['continent'], match['region']))

SQLITE3_DB.commit()

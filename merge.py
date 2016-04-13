#!/usr/bin/env python3
# encoding=utf8
# 指定一个目录，将该目录下所有xlsx文件和data_result.sqlite3中的数据合并


import sys
import os
import sqlite3
from openpyxl import load_workbook

conn = sqlite3.connect('./data_result.sqlite3')
cursor = conn.cursor()

count = 0

for parent, dirnames, filenames in os.walk(sys.argv[1]):
    for filename in filenames:
        if not filename.endswith('.xlsx'):
            continue
        path = parent + '/' + filename

        wb = load_workbook(path)
        ws = wb.active
        row = 1
        while ws['A' + str(row)].value:
            rid = str(row)
            year = ws['J' + rid].value
            name = ws['A' + rid].value
            if name.startswith('.'):
                name = name[1:]
            cursor.execute(
                'select commet,address,content,comment_todo,content_todo from data where year = ? and filename = ?', (year, name))
            data = cursor.fetchall()
            if data:
                count += 1
                data = data[0]
                ws['K' + rid] = data[0]
                ws['L' + rid] = data[1]
                ws['M' + rid] = data[2]
                ws['N' + rid] = data[3]
                ws['O' + rid] = data[4]
            else:
                placeholder = '***********************'
                ws['k' + rid] = placeholder
                ws['L' + rid] = placeholder
                ws['M' + rid] = placeholder
                ws['N' + rid] = placeholder
                ws['O' + rid] = placeholder
                print(filename, name)
            row += 1
        wb.save('./data/' + filename)

print(count)
cursor.close()

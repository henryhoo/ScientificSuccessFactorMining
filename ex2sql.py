#!/usr/bin/env python3
# encoding=utf8

from openpyxl import load_workbook
import sqlite3
import sys
import os


SQLITE3_FILE = './data.sqlite3'

try:
    os.unlink(SQLITE3_FILE)
except:
    pass

conn = sqlite3.connect(SQLITE3_FILE)

print('Be Patient!')

try:
    c = conn.cursor()
    c.execute(
        'create table data(year text,filename text, author text,article text, journal text, page text, publish text, comment text, address text, content text,reference text)')
    conn.commit()
    c.close()
except:
    pass


def insert_data(year, filename, author, article, journal,
                page, publish, comment, address, content, reference):
    global conn
    c = conn.cursor()
    c.execute('insert into data values(?,?,?,?,?,?,?,?,?,?,?)',
              (str(year), str(filename), str(author), str(article), str(journal),
               str(page), str(publish), str(comment), str(address),
               str(content), str(reference)))
    conn.commit()
    c.close()

if len(sys.argv) != 2:
    sys.exit(-1)

dir = sys.argv[1]


for parent, dirnames, filenames in os.walk(dir):
    for filename in filenames:
        if filename.endswith('xlsx'):
            path = parent + '/' + filename
            try:
                wb = load_workbook(path)
                ws = wb['Sheet1']
                row = 1
                while ws['A' + str(row)].value:
                    rid = str(row)
                    insert_data(ws['A' + rid].value, ws['B' + rid].value,
                                ws['C' + rid].value, ws['D' + rid].value,
                                ws['E' + rid].value, ws['F' + rid].value,
                                ws['G' + rid].value, ws['H' + rid].value,
                                ws['I' + rid].value, ws['J' + rid].value,
                                ws['K' + rid].value)
                    row += 1
            except Exception as e:
                raise e
                print(e)

#!/usr/bin/env python3
# encoding=utf8

# 将sqlite3格式的数据转化为excel


import sqlite3
import sys
import os
import openpyxl


if len(sys.argv) != 2:
    print('%s topic.sqlite3' % sys.argv[0])
    sys.exit(0)

TOPIC_FILE = sys.argv[1]
OUTPUT = 'topic.xlsx'

try:
    os.unlink(OUTPUT)
except:
    pass


wb = openpyxl.Workbook()
wb.remove_sheet(wb.active)
TOPIC_DB = sqlite3.connect(TOPIC_FILE)
cur = TOPIC_DB.cursor()

# 主题词表
ws = wb.create_sheet()
ws.title = '主题词表'
cur.execute('select topic, phrase,M,P,Q from topic')
rows = cur.fetchall()

i = 1
for row in rows:
    ws['A' + str(i)] = row[0]
    ws['B' + str(i)] = row[1]
    ws['C' + str(i)] = row[2]
    ws['D' + str(i)] = row[3]
    ws['E' + str(i)] = row[4]
    i += 1

# topics
cur.execute('select topic from basic group by topic')
rows = cur.fetchall()

TOPICS = []
for row in rows:
    TOPICS.append(row[0])

# state_year表
ws = wb.create_sheet()
ws.title = 'state_year表'
cur.execute('select state_year from state_year group by '
            'state_year order by state_year')
rows = cur.fetchall()
YEARS = {}

col = ord('B')
for row in rows:
    ws[chr(col) + '1'] = row[0]
    YEARS[row[0]] = chr(col)
    col += 1
ws[chr(col) + '1'] = 'total'
YEARS['total'] = chr(col)


i = 2
for tp in TOPICS:
    ws['A'+str(i)] = tp
    cur.execute('select state_year,M from state_year where topic=?', (tp,))
    rows = cur.fetchall()
    total = 0
    for row in rows:
        year = row[0]
        M = row[1]
        total += M
        ws[YEARS[year] + str(i)] = M
    ws[YEARS['total'] + str(i)] = total
    i += 1
# state_year_detail表
ws = wb.create_sheet()
ws.title = 'state_year_detail表'
cur.execute('select state_year from state_year group by '
            'state_year order by state_year')
rows = cur.fetchall()
YEARS = {}

col = ord('C')
for row in rows:
    ws[chr(col) + '1'] = row[0]
    YEARS[row[0]] = chr(col)
    col += 1
ws[chr(col) + '1'] = 'total'
YEARS['total'] = chr(col)

i = 2
for tp in TOPICS:
    ws['A' + str(i)] = tp
    cur.execute('select phrase from state_year_detail where topic=? '
                'group by phrase', (tp,))
    phrases = cur.fetchall()
    for row in phrases:
        phrase = row[0]
        cur.execute('select phrase, state_year,N from state_year_detail where '
                    'topic=? and phrase=?', (tp, phrase))
        rows = cur.fetchall()
        total = 0
        ws['A' + str(i)] = tp
        ws['B' + str(i)] = phrase
        for row in rows:
            state_year = row[1]
            N = row[2]
            total += N
            ws[YEARS[state_year] + str(i)] = N
        ws[YEARS['total'] + str(i)] = total
        i += 1

# continent表
ws = wb.create_sheet()
ws.title = 'continent表'
cur.execute('select continent from continent group by '
            'continent order by continent')
rows = cur.fetchall()
CONTINENTS = {}

col = ord('B')
for row in rows:
    ws[chr(col) + '1'] = row[0]
    CONTINENTS[row[0]] = chr(col)
    col += 1
ws[chr(col) + '1'] = 'total'
CONTINENTS['total'] = chr(col)


i = 2
for tp in TOPICS:
    ws['A'+str(i)] = tp
    cur.execute('select continent,M from continent where topic=?', (tp,))
    rows = cur.fetchall()
    total = 0
    for row in rows:
        continent = row[0]
        M = row[1]
        total += M
        ws[CONTINENTS[continent] + str(i)] = M
    ws[CONTINENTS['total'] + str(i)] = total
    i += 1

# continent_detail表
ws = wb.create_sheet()
ws.title = 'continent_detail表'
cur.execute('select continent from continent group by '
            'continent order by continent')
rows = cur.fetchall()
CONTINENTS = {}

col = ord('C')
for row in rows:
    ws[chr(col) + '1'] = row[0]
    CONTINENTS[row[0]] = chr(col)
    col += 1
ws[chr(col) + '1'] = 'total'
CONTINENTS['total'] = chr(col)

i = 2
for tp in TOPICS:
    ws['A' + str(i)] = tp
    cur.execute('select phrase from continent_detail where topic=? '
                'group by phrase', (tp,))
    phrases = cur.fetchall()
    for row in phrases:
        phrase = row[0]
        cur.execute('select phrase, continent,N from continent_detail where '
                    'topic=? and phrase=?', (tp, phrase))
        rows = cur.fetchall()
        total = 0
        ws['A' + str(i)] = tp
        ws['B' + str(i)] = phrase
        for row in rows:
            continent = row[1]
            N = row[2]
            total += N
            ws[CONTINENTS[continent] + str(i)] = N
        ws[CONTINENTS['total'] + str(i)] = total
        i += 1

# region表
ws = wb.create_sheet()
ws.title = 'region表'
cur.execute('select region from region group by '
            'region order by region')
rows = cur.fetchall()
REGIONS = {}

col = ord('B')
for row in rows:
    ws[chr(col) + '1'] = row[0]
    REGIONS[row[0]] = chr(col)
    col += 1
ws[chr(col) + '1'] = 'total'
REGIONS['total'] = chr(col)


i = 2
for tp in TOPICS:
    ws['A'+str(i)] = tp
    cur.execute('select region,M from region where topic=?', (tp,))
    rows = cur.fetchall()
    total = 0
    for row in rows:
        region = row[0]
        M = row[1]
        total += M
        ws[REGIONS[region] + str(i)] = M
    ws[REGIONS['total'] + str(i)] = total
    i += 1

# region_detail表
ws = wb.create_sheet()
ws.title = 'region_detail表'
cur.execute('select region from region group by '
            'region order by region')
rows = cur.fetchall()
REGIONS = {}

col = ord('C')
for row in rows:
    ws[chr(col) + '1'] = row[0]
    REGIONS[row[0]] = chr(col)
    col += 1
ws[chr(col) + '1'] = 'total'
REGIONS['total'] = chr(col)

i = 2
for tp in TOPICS:
    ws['A' + str(i)] = tp
    cur.execute('select phrase from region_detail where topic=? '
                'group by phrase', (tp,))
    phrases = cur.fetchall()
    for row in phrases:
        phrase = row[0]
        cur.execute('select phrase, region,N from region_detail where '
                    'topic=? and phrase=?', (tp, phrase))
        rows = cur.fetchall()
        total = 0
        ws['A' + str(i)] = tp
        ws['B' + str(i)] = phrase
        for row in rows:
            region = row[1]
            N = row[2]
            total += N
            ws[REGIONS[region] + str(i)] = N
        ws[REGIONS['total'] + str(i)] = total
        i += 1

# year 表
ws = wb.create_sheet()
ws.title = 'year表'

ws['B1'] = '1940以前'
ws['C1'] = '1940-1949'
ws['D1'] = '1950-1959'
ws['E1'] = '1960-1969'
ws['F1'] = '1970-1979'
ws['G1'] = '1980以后'


def get_year_col(year):
    if year < 1940:
        return 'B'
    elif year >= 1940 and year <= 1949:
        return 'C'
    elif year >= 1950 and year <= 1959:
        return 'D'
    elif year >= 1960 and year <= 1969:
        return 'E'
    elif year >= 1970 and year <= 1979:
        return 'F'
    return 'G'

i = 2
for tp in TOPICS:
    ws['A'+str(i)] = tp
    cur.execute('select year,M from year where topic=?', (tp,))
    rows = cur.fetchall()
    for row in rows:
        year = row[0]
        M = row[1]
        col = get_year_col(int(year)) + str(i)
        if ws[col].value:
            ws[col] = ws[col].value + M
        else:
            ws[col] = M
    i += 1

# year_detail 表
ws = wb.create_sheet()
ws.title = 'year_detail表'

ws['C1'] = '1940以前'
ws['D1'] = '1940-1949'
ws['E1'] = '1950-1959'
ws['F1'] = '1960-1969'
ws['G1'] = '1970-1979'
ws['H1'] = '1980以后'


def get_year_col(year):
    if year < 1940:
        return 'C'
    elif year >= 1940 and year <= 1949:
        return 'D'
    elif year >= 1950 and year <= 1959:
        return 'E'
    elif year >= 1960 and year <= 1969:
        return 'F'
    elif year >= 1970 and year <= 1979:
        return 'G'
    return 'H'

i = 2
for tp in TOPICS:
    ws['A'+str(i)] = tp
    cur.execute('select phrase from year_detail where topic=? '
                'group by phrase;', (tp,))
    phrases = cur.fetchall()
    for row in phrases:
        phrase = row[0]
        cur.execute('select phrase,year,N from year_detail where topic=? '
                    'and phrase=?', (tp, phrase))
        rows = cur.fetchall()
        ws['A' + str(i)] = tp
        ws['B' + str(i)] = phrase
        for row in rows:
            year = row[1]
            N = row[2]
            col = get_year_col(int(year)) + str(i)
            if ws[col].value:
                ws[col] = ws[col].value + N
            else:
                ws[col] = N
        i += 1

# type 表
ws = wb.create_sheet()
ws.title = 'type表'

cur.execute('select type from type group by type')
rows = cur.fetchall()

TYPES = {}
col = ord('B')
for row in rows:
    ws[chr(col) + '1'] = row[0]
    TYPES[row[0]] = chr(col)
    col += 1

i = 2
for tp in TOPICS:
    ws['A'+str(i)] = tp
    cur.execute('select type, M from type where topic=?', (tp,))
    rows = cur.fetchall()
    for row in rows:
        t = row[0]
        M = row[1]
        col = TYPES[t] + str(i)
        ws[col] = M
    i += 1

# type_detail表
ws = wb.create_sheet()
ws.title = 'type_detail表'

cur.execute('select type from type_detail group by type')
rows = cur.fetchall()

TYPES = {}
col = ord('C')
for row in rows:
    ws[chr(col) + '1'] = row[0]
    TYPES[row[0]] = chr(col)
    col += 1

i = 2
for tp in TOPICS:
    cur.execute('select phrase from type_detail where topic=? group by phrase',
                (tp,))
    phrases = cur.fetchall()
    for row in phrases:
        phrase = row[0]
        ws['A' + str(i)] = tp
        ws['B' + str(i)] = phrase
        cur.execute('select phrase, type, N from type_detail where topic = ?'
                    ' and phrase=?', (tp, phrase))
        rows = cur.fetchall()
        for row in rows:
            t = row[1]
            N = row[2]
            ws[TYPES[t] + str(i)] = N
        i += 1

wb.save(OUTPUT)
